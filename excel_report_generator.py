import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def load_all_sales(folder_path):
    folder = Path(folder_path)
    all_dataframes = []

    for file in folder.glob("ventes_*.xlsx"):
        df = pd.read_excel(file)
        all_dataframes.append(df)
        print(f"Charge : {file.name} ({len(df)} lignes)")

    combined = pd.concat(all_dataframes, ignore_index=True)
    return combined


def consolidate_by_product(df):
    consolidated = df.groupby("Produit").agg({
        "Quantite": "sum",
        "Prix_unitaire": "mean",
    }).reset_index()

    return consolidated


def format_report(filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = max_length + 4

    workbook.save(filepath)


if __name__ == "__main__":
    result = load_all_sales(".")
    consolidated = consolidate_by_product(result)
    print(consolidated)

    output_file = "rapport_consolide.xlsx"
    consolidated.to_excel(output_file, index=False)
    format_report(output_file)
    print(f"Rapport exporte et mis en forme : {output_file}")